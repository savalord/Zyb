import os
import numpy
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import defcalculate as df
from openpyxl.styles import Alignment
import fio as f
import re
from os import path




misum = []
masum = []
def upload_excel(filename):
    global masum
    global misum
    wb = load_workbook(filename)
    sheet4 = wb['Лист1']
    sheet2 = wb['2']

    df2 = pd.DataFrame(sheet2.values)
    #df2.columns = ["0", "1", "2", "3", "4"]
    #df2 = df2.drop(columns=['2','3','4'])
    df2.columns = ["ФИО","Класс"]
    #df2 = df2.drop(columns=["0", "1", "2", "3", "4","5", "6", "7"])
    df4 = pd.DataFrame(sheet4.values)
    df3 = pd.DataFrame(sheet4.values)
    df5 = pd.DataFrame(sheet4.values)
    df6 = pd.DataFrame(sheet4.values)
    df4.columns = ["ID", "KOD", "NAME", "TARIF", "COUNT", "SUMM"]
    df3.columns = ["ID", "KOD", "NAME", "TARIF", "COUNT", "SUMM"]
    df5.columns = ["ID", "KOD", "NAME", "TARIF", "COUNT", "SUMM"]
    df6.columns = ["ID", "KOD", "NAME", "TARIF", "COUNT", "SUMM"]
    df5 = df5.drop(columns=["ID","KOD", "NAME", "TARIF","SUMM"])
    df6 = df6.drop(columns=["ID", "KOD", "NAME", "TARIF", "SUMM"])
    #df5["COUNT"] = 0
    #df6["COUNT"] = 0
    #fgyt = df5["COUNT"] # ДЛЯ подсчета кодов
    #fgyt12 = df6["COUNT"] # ДЛЯ подсчета кодов

    fgclass = df2["Класс"]
    fg1 = df3["ID"]
    fg2 = df3["KOD"]
    fg3 = df3["NAME"]
    fg4 = df3["TARIF"]
    fgcount = df3["COUNT"]
    fg6 = df3["SUMM"]
    fg4 = fg4.ffill()

    fgcount = fgcount.array
    fgcount = fgcount.to_numpy()

    fgclass = fgclass.array
    fgclass = fgclass.to_numpy()

    fg2 = fg2.array
    fg2 = fg2.to_numpy()


    fg4 = fg4.array
    fg4 = fg4[numpy.logical_not(numpy.isnan(fg4))]
    fg4 = fg4.to_numpy()

    df4 = df4.drop('COUNT', axis=1)
    df4 = df4.drop('SUMM', axis=1)
    df4 = df4.drop('NAME', axis=1)


    localcountmin, localcountmax = df.localcount(fgcount,fgclass)
    localsummin, localsummax = df.local(fgclass)
    fioiof = df2["ФИО"]
    Sminmax = df.calculate(fgclass,fg2,fg4,localsummin,localsummax,fg3,localcountmin, fioiof,localcountmax)
    Df_Sminmax = pd.DataFrame(Sminmax)


    masum = localsummax
    misum = localsummin



def save_excel():
    global masum
    global misum
    filename = './Exsel_Files/'
    df2 = pd.DataFrame({'Минимальная цена': [df.summin], 'Оптимальная цена': [df.summax]})

    #df2.to_excel(writer, sheet_name='Лист1', startrow=1, index=False)
    #df.Names = df.Names.fillna(method='ffill')



    # добавляем данные в активный лист

    #file_name = filedialog.askopenfilename(initialdir=path.dirname(__file__))  # получаем путь нужный файл
    r = filename.rfind("/")  # находит последний знак и возвращает его индекс


    # сохраняем

    rt = df.fio
    print(rt)
    #rt = rt.values.tolist()

    rt = rt.replace(r'\s+','',regex=True)
    test = len(rt)
    print(rt)
    rtstring = []
    for i in range(len(rt)):
        temp = rt.iloc[i]
        print(temp.dtypes)
        rtstring.append(str(temp))
    #rt = str(rt)# ошибка здесь кушает фамилии
    print(rtstring)

    fio_list = f.format_fio(rtstring)
    fio_list = re.sub("[0-9],", "", str(fio_list))
    fio_list = re.sub("ФИО", "", str(fio_list))
    fio_list = re.sub("Name:", " ", str(fio_list))
    fio_list = re.sub("dtype:object", " ", str(fio_list))
    symbols_to_remove = ",[]'"

    for symbol in symbols_to_remove:
        fio_list = fio_list.replace(symbol, "")
    fio_list = re.sub("[0-9]", "", str(fio_list))
    print(fio_list)
    fio_list = fio_list.split()
    print("fio_list[1]")
    print(fio_list[1])
    print(fio_list)
    x = 0
    pdmasum = pd.DataFrame(masum)
    pdmasum.columns = ['Макс']
    pdmisum = pd.DataFrame(misum)
    pdmisum.columns = ['Мин']
    print(pdmasum)
    for i in range(0, len(fio_list)):
    #for i in range(0, 10):
        x = x + 1
        my_file_name = filename[:r]  # сохраняем путь включая этот индекс. т.е получаем нашу дирректорию

        writer = pd.ExcelWriter(my_file_name + '/' + fio_list[i] + 'Мин.xlsx')
        pdmisum[i:i+1].to_excel(writer, sheet_name='Лист1', startrow=84, startcol=8, index=False)
        df.Names.to_excel(writer, sheet_name='Лист1', startrow=1, startcol=2, index=False)
        df.dfert.to_excel(writer, sheet_name='Лист1', startrow=1, index=False)
        df.TariFF.to_excel(writer, sheet_name='Лист1', startrow=1, startcol=8, index=False)
        df.mincount[i].to_excel(writer, sheet_name='Лист1', startrow=1, startcol=9, index=False)
        writer.close()
        wb = openpyxl.load_workbook(my_file_name + '/' + fio_list[i] + 'Мин.xlsx')
        ws = wb.active
        ws['A1'] = fio_list[i]
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['A2'] = 'Код'
        ws['C2'] = 'Наименование услуги'
        ws['I2'] = 'Цена'
        ws['J2'] = 'Количество'
        ws['G86'] = 'Итоговая Сумма'
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
        ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=8)
        # НАЗВАНИЯ
        ws.merge_cells(start_row=13, start_column=3, end_row=16, end_column=8)
        ws.merge_cells(start_row=21, start_column=3, end_row=24, end_column=8)
        ws.merge_cells(start_row=27, start_column=3, end_row=30, end_column=8)
        ws.merge_cells(start_row=31, start_column=3, end_row=34, end_column=8)
        ws.merge_cells(start_row=35, start_column=3, end_row=38, end_column=8)
        ws.merge_cells(start_row=39, start_column=3, end_row=42, end_column=8)
        ws.merge_cells(start_row=43, start_column=3, end_row=46, end_column=8)
        ws.merge_cells(start_row=47, start_column=3, end_row=50, end_column=8)
        ws.merge_cells(start_row=52, start_column=3, end_row=55, end_column=8)
        ws.merge_cells(start_row=62, start_column=3, end_row=67, end_column=8)

        # Цены
        ws.merge_cells(start_row=13, start_column=9, end_row=16, end_column=9)
        ws.merge_cells(start_row=21, start_column=9, end_row=24, end_column=9)
        ws.merge_cells(start_row=27, start_column=9, end_row=30, end_column=9)
        ws.merge_cells(start_row=31, start_column=9, end_row=34, end_column=9)
        ws.merge_cells(start_row=35, start_column=9, end_row=38, end_column=9)
        ws.merge_cells(start_row=39, start_column=9, end_row=42, end_column=9)
        ws.merge_cells(start_row=43, start_column=9, end_row=46, end_column=9)
        ws.merge_cells(start_row=47, start_column=9, end_row=50, end_column=9)
        ws.merge_cells(start_row=52, start_column=9, end_row=55, end_column=9)
        ws.merge_cells(start_row=62, start_column=9, end_row=67, end_column=9)

        # Количесво
        ws.merge_cells(start_row=13, start_column=10, end_row=16, end_column=10)
        ws.merge_cells(start_row=21, start_column=10, end_row=24, end_column=10)
        ws.merge_cells(start_row=27, start_column=10, end_row=30, end_column=10)
        ws.merge_cells(start_row=31, start_column=10, end_row=34, end_column=10)
        ws.merge_cells(start_row=35, start_column=10, end_row=38, end_column=10)
        ws.merge_cells(start_row=39, start_column=10, end_row=42, end_column=10)
        ws.merge_cells(start_row=43, start_column=10, end_row=46, end_column=10)
        ws.merge_cells(start_row=47, start_column=10, end_row=50, end_column=10)
        ws.merge_cells(start_row=52, start_column=10, end_row=55, end_column=10)
        ws.merge_cells(start_row=62, start_column=10, end_row=67, end_column=10)

        wb.save(my_file_name + '/' + fio_list[i] + 'Мин.xlsx')

        writer = pd.ExcelWriter(my_file_name + '/' + fio_list[i] + 'Макс.xlsx')
        pdmasum[i:i+1].to_excel(writer, sheet_name='Лист1', startrow=84, startcol=8, index=False)
        df.Names.to_excel(writer, sheet_name='Лист1', startrow=1, startcol=2, index=False)
        df.dfert.to_excel(writer, sheet_name='Лист1', startrow=1, index=False)
        df.TariFF.to_excel(writer, sheet_name='Лист1', startrow=1, startcol=8, index=False)
        df.maxcount[i].to_excel(writer, sheet_name='Лист1', startrow=1, startcol=9, index=False)
        writer.close()
        wb = openpyxl.load_workbook(my_file_name + '/' + fio_list[i] + 'Макс.xlsx')
        ws = wb.active
        ws['A1'] = fio_list[i]
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['G86'] ='Итоговая Сумма'
        ws['A2'] = 'Код'
        ws['C2'] = 'Наименование услуги'
        ws['I2'] = 'Цена'
        ws['J2'] = 'Количество'
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
        ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=8)
        # НАЗВАНИЯ
        ws.merge_cells(start_row=13, start_column=3, end_row=16, end_column=8)
        ws.merge_cells(start_row=21, start_column=3, end_row=24, end_column=8)
        ws.merge_cells(start_row=27, start_column=3, end_row=30, end_column=8)
        ws.merge_cells(start_row=31, start_column=3, end_row=34, end_column=8)
        ws.merge_cells(start_row=35, start_column=3, end_row=38, end_column=8)
        ws.merge_cells(start_row=39, start_column=3, end_row=42, end_column=8)
        ws.merge_cells(start_row=43, start_column=3, end_row=46, end_column=8)
        ws.merge_cells(start_row=47, start_column=3, end_row=50, end_column=8)
        ws.merge_cells(start_row=52, start_column=3, end_row=55, end_column=8)
        ws.merge_cells(start_row=62, start_column=3, end_row=67, end_column=8)

        # Цены
        ws.merge_cells(start_row=13, start_column=9, end_row=16, end_column=9)
        ws.merge_cells(start_row=21, start_column=9, end_row=24, end_column=9)
        ws.merge_cells(start_row=27, start_column=9, end_row=30, end_column=9)
        ws.merge_cells(start_row=31, start_column=9, end_row=34, end_column=9)
        ws.merge_cells(start_row=35, start_column=9, end_row=38, end_column=9)
        ws.merge_cells(start_row=39, start_column=9, end_row=42, end_column=9)
        ws.merge_cells(start_row=43, start_column=9, end_row=46, end_column=9)
        ws.merge_cells(start_row=47, start_column=9, end_row=50, end_column=9)
        ws.merge_cells(start_row=52, start_column=9, end_row=55, end_column=9)
        ws.merge_cells(start_row=62, start_column=9, end_row=67, end_column=9)

        # Количесво
        ws.merge_cells(start_row=13, start_column=10, end_row=16, end_column=10)
        ws.merge_cells(start_row=21, start_column=10, end_row=24, end_column=10)
        ws.merge_cells(start_row=27, start_column=10, end_row=30, end_column=10)
        ws.merge_cells(start_row=31, start_column=10, end_row=34, end_column=10)
        ws.merge_cells(start_row=35, start_column=10, end_row=38, end_column=10)
        ws.merge_cells(start_row=39, start_column=10, end_row=42, end_column=10)
        ws.merge_cells(start_row=43, start_column=10, end_row=46, end_column=10)
        ws.merge_cells(start_row=47, start_column=10, end_row=50, end_column=10)
        ws.merge_cells(start_row=52, start_column=10, end_row=55, end_column=10)
        ws.merge_cells(start_row=62, start_column=10, end_row=67, end_column=10)

        wb.save(my_file_name + '/' +  fio_list[i] + 'Макс.xlsx')
    web = Workbook()
    ws = web.active
    ws['A1'] = 'Максимальная сумма'
    ws['B1'] = 'Минимальная сумма'
    ws['A2'] = df.summax
    ws['B2'] = df.summin
    my_file_name = filename[:r]
    web.save(my_file_name + '/' + 'общая_сумма.xlsx')
    print("Готово")

    return fio_list
